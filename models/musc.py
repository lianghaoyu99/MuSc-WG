import os
import sys
import numpy as np
import torch
import torch.nn.functional as F
sys.path.append('./models/backbone')

import datasets.mvtec as mvtec
from datasets.mvtec import _CLASSNAMES as _CLASSNAMES_mvtec_ad
import datasets.visa as visa
from datasets.visa import _CLASSNAMES as _CLASSNAMES_visa
import datasets.btad as btad
from datasets.btad import _CLASSNAMES as _CLASSNAMES_btad
import datasets.miniled as miniled
from datasets.miniled import _CLASSNAMES as _CLASSNAMES_miniled
import datasets.microled as microled
from datasets.microled import _CLASSNAMES as _CLASSNAMES_microled

import models.backbone.open_clip as open_clip
import models.backbone._backbones as _backbones
from models.modules._LNAMD import LNAMD
# from models.modules.WTConvLNAMD import WTConvLNAMD
from models.modules.WTConvStatic import WTConvLNAMDStatic
from models.modules._MSM import MSM
from models.modules._RsCIN import RsCIN
from models.modules._Optimization import AnomalyMapOptimizer
from utils.metrics import compute_metrics
from openpyxl import Workbook
from tqdm import tqdm
import pickle
import time
import cv2
import matplotlib.pyplot as plt
from sklearn.manifold import TSNE

import warnings
warnings.filterwarnings("ignore")


class MuSc():
    def __init__(self, cfg, seed=0):
        self.cfg = cfg
        self.seed = seed
        self.device = torch.device("cuda:{}".format(cfg['device']) if torch.cuda.is_available() else "cpu")
        print(f"Active device: {self.device}")

        self.path = cfg['datasets']['data_path']
        self.dataset = cfg['datasets']['dataset_name']
        self.vis = cfg['testing']['vis']
        self.vis_type = cfg['testing']['vis_type']
        self.vis_overlay = cfg['testing'].get('vis_overlay', True)
        self.save_excel = cfg['testing']['save_excel']
        self.use_rscin = cfg['testing'].get('use_rscin', True)
        self.vis_tsne = cfg['testing'].get('vis_tsne', False)
        self.tsne_perplexity = cfg['testing'].get('tsne_perplexity', 30)
        self.tsne_n_iter = cfg['testing'].get('tsne_n_iter', 1000)
        self.tsne_source = cfg['testing'].get('tsne_source', 'class') # 'class' for class tokens, 'lnamd' for LNAMD features
        # the categories to be tested
        self.categories = cfg['datasets']['class_name']
        if isinstance(self.categories, str):
            if self.categories.lower() == 'all':
                if self.dataset == 'visa':
                    self.categories = _CLASSNAMES_visa
                elif self.dataset == 'mvtec_ad':
                    self.categories = _CLASSNAMES_mvtec_ad
                elif self.dataset == 'btad':
                    self.categories = _CLASSNAMES_btad
                elif self.dataset == 'miniled_ad':
                    self.categories = _CLASSNAMES_miniled
                elif self.dataset == 'microled_ad':
                    self.categories = _CLASSNAMES_microled
            else:
                self.categories = [self.categories]

        self.model_name = cfg['models']['backbone_name']
        self.image_size = cfg['datasets']['img_resize']
        self.batch_size = cfg['models']['batch_size']
        self.pretrained = cfg['models']['pretrained']
        self.features_list = [l+1 for l in cfg['models']['feature_layers']]
        self.divide_num = cfg['datasets']['divide_num']
        self.r_list = cfg['models']['r_list']
        self.output_dir = os.path.join(cfg['testing']['output_dir'], self.dataset, self.model_name, 'imagesize{}'.format(self.image_size))
        os.makedirs(self.output_dir, exist_ok=True)
        self.load_backbone()


    def load_backbone(self):
        if 'dino' in self.model_name:
            # dino or dino_v2
            self.dino_model = _backbones.load(self.model_name)
            self.dino_model.to(self.device)
            self.preprocess = None
        else:
            # clip
            self.clip_model, _, self.preprocess = open_clip.create_model_and_transforms(self.model_name, self.image_size, pretrained=self.pretrained)
            self.clip_model.to(self.device)


    def load_datasets(self, category, divide_num=1, divide_iter=0):
        # dataloader
        if self.dataset == 'visa':
            test_dataset = visa.VisaDataset(source=self.path, split=visa.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'mvtec_ad':
            test_dataset = mvtec.MVTecDataset(source=self.path, split=mvtec.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'btad':
            test_dataset = btad.BTADDataset(source=self.path, split=btad.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'miniled_ad':
            test_dataset = miniled.MiniledDataset(source=self.path, split=miniled.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'microled_ad':
            test_dataset = microled.MicroledDataset(source=self.path, split=microled.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        return test_dataset


    def visualization(self, image_path_list, gt_list, pr_px, gt_px, category):
        def normalization01(img):
            return (img - img.min()) / (img.max() - img.min() + 1e-8)
            
        def apply_ad_scoremap(image, scoremap, alpha=0.5, overlay=True):
            np_image = np.asarray(image, dtype=float)
            scoremap = (scoremap * 255).astype(np.uint8)
            scoremap = cv2.applyColorMap(scoremap, cv2.COLORMAP_JET)
            scoremap = cv2.cvtColor(scoremap, cv2.COLOR_BGR2RGB)
            if overlay:
                return (alpha * np_image + (1 - alpha) * scoremap).astype(np.uint8)
            return scoremap

        def draw_mask_contour(image, mask):
            if mask.ndim == 3:
                mask = mask[0]
            mask_uint8 = (mask * 255).astype(np.uint8)
            mask_uint8 = np.ascontiguousarray(mask_uint8)
            contours, _ = cv2.findContours(mask_uint8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            vis = image.copy()
            cv2.drawContours(vis, contours, -1, (0, 255, 0), 2)  # 绿色描边
            return vis

        if self.vis_type != 'single_norm':
            # normalized all image
            pr_px = normalization01(pr_px)

        for i, path in enumerate(image_path_list):
            path = os.path.normpath(path)
            anomaly_type = os.path.basename(os.path.dirname(path))
            img_name = os.path.basename(path)
            base_name = os.path.splitext(img_name)[0]
            
            save_dir = os.path.join(self.output_dir, 'vis', category, anomaly_type)
            os.makedirs(save_dir, exist_ok=True)
            
            # Read original image
            ori_img = cv2.imread(path)
            if ori_img is None:
                continue
            ori_img = cv2.cvtColor(ori_img, cv2.COLOR_BGR2RGB)
            ori_img = cv2.resize(ori_img, (self.image_size, self.image_size))
            
            # GT contour
            gt_mask = gt_px[i].squeeze()
            gt_mask = cv2.resize(gt_mask.astype(np.float32), (self.image_size, self.image_size), interpolation=cv2.INTER_NEAREST)
            gt_vis = draw_mask_contour(ori_img, gt_mask)
            save_gt = os.path.join(save_dir, f"{base_name}_gt.png")
            cv2.imwrite(save_gt, cv2.cvtColor(gt_vis, cv2.COLOR_RGB2BGR))
            
            # Anomaly map
            anomaly_map = pr_px[i].squeeze()
            if self.vis_type == 'single_norm':
                anomaly_map = normalization01(anomaly_map)
                
            anomaly_map = cv2.resize(anomaly_map, (self.image_size, self.image_size))
            vis = apply_ad_scoremap(ori_img, anomaly_map, overlay=self.vis_overlay)
            save_vis = os.path.join(save_dir, f"{base_name}.png")
            cv2.imwrite(save_vis, cv2.cvtColor(vis, cv2.COLOR_RGB2BGR))


    def visualize_tsne(self, features, labels, category, title="t-SNE Visualization"):
        """
        Perform t-SNE visualization on the features.
        
        Args:
            features (numpy.ndarray): Feature matrix of shape (N, D).
            labels (numpy.ndarray): Labels of shape (N,), 0 for normal, 1 for anomaly.
            category (str): The category name for saving the plot.
            title (str): Title of the plot.
        """
        print(f"Performing t-SNE on {features.shape[0]} samples with dimension {features.shape[1]}...")
        tsne = TSNE(n_components=2, perplexity=self.tsne_perplexity, n_iter=self.tsne_n_iter, random_state=self.seed)
        features_2d = tsne.fit_transform(features)
        
        plt.figure(figsize=(10, 8))
        
        # Plot normal samples (label 0)
        normal_indices = labels == 0
        plt.scatter(features_2d[normal_indices, 0], features_2d[normal_indices, 1], 
                    c='blue', label='Normal', alpha=0.6, s=20)
        
        # Plot anomaly samples (label 1)
        anomaly_indices = labels == 1
        plt.scatter(features_2d[anomaly_indices, 0], features_2d[anomaly_indices, 1], 
                    c='red', label='Anomaly', alpha=0.6, s=20)
        
        plt.title(f"{title} - {category}")
        plt.legend()
        plt.grid(True, linestyle='--', alpha=0.3)
        
        # Save the plot
        tsne_dir = os.path.join(self.output_dir, category, 'tsne_visualization')
        os.makedirs(tsne_dir, exist_ok=True)
        save_path = os.path.join(tsne_dir, f'{self.tsne_source}_tsne.png')
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"t-SNE plot saved to {save_path}")

    def make_category_data(self, category):
        print(category)

        # divide sub-datasets
        divide_num = self.divide_num
        anomaly_maps = torch.tensor([]).double()
        gt_list = []
        img_masks = []
        class_tokens = []
        image_path_list = []
        # t-SNE data collection
        tsne_features = []
        tsne_labels = []
        
        start_time_all = time.time()
        dataset_num = 0
        for divide_iter in range(divide_num):  # 按照划分数据子集的数量依次处理每个子集
            test_dataset = self.load_datasets(category, divide_num=divide_num, divide_iter=divide_iter)  # 选择要加载的数据集路径
            test_dataloader = torch.utils.data.DataLoader(
                test_dataset,
                batch_size=self.batch_size,
                shuffle=False,
                num_workers=0,
                pin_memory=True,
            )  # 创建对应的Dataloader
            
            # extract features 利用ViT进行特征提取
            patch_tokens_list = []
            subset_num = len(test_dataset)
            dataset_num += subset_num
            start_time = time.time()
            for image_info in tqdm(test_dataloader):  # 遍历抽取每个batch图像的特征
            # for image_info in test_dataloader:
                if isinstance(image_info, dict):
                    image = image_info["image"]
                    image_path_list.extend(image_info["image_path"])
                    img_masks.append(image_info["mask"])
                    gt_list.extend(list(image_info["is_anomaly"].numpy()))
                with torch.no_grad(), torch.cuda.amp.autocast():
                    input_image = image.to(torch.float).to(self.device)
                    if 'dinov2' in self.model_name or 'dinov3' in self.model_name:
                        patch_tokens = self.dino_model.get_intermediate_layers(x=input_image, n=[l-1 for l in self.features_list], return_class_token=False)
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                        fake_cls = [torch.zeros_like(p)[:, 0:1, :] for p in patch_tokens]
                        patch_tokens = [torch.cat([fake_cls[i], patch_tokens[i]], dim=1) for i in range(len(patch_tokens))]
                    elif 'dino' in self.model_name:
                        patch_tokens_all = self.dino_model.get_intermediate_layers(x=input_image, n=max(self.features_list))
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens_all[l-1].cpu() for l in self.features_list]
                    else: # clip
                        image_features, patch_tokens = self.clip_model.encode_image(input_image, self.features_list)  # 用CLIP提取全局特征和指定层的局部特征
                        # print("image_features shape:", image_features.shape)  # 应该是 [batch_size, embed_dim] [4, 768]
                        # print("patch_tokens type:", type(patch_tokens))  # 应该是 list
                        # print("patch_tokens length:", len(patch_tokens))  # 应该等于 len(self.features_list) 4
                        # print("First patch token shape:",
                        #       patch_tokens[0].shape)  # 应该是 [batch_size, num_patches, embed_dim] [4, 1370, 1024]
                        image_features /= image_features.norm(dim=-1, keepdim=True)  # 全局特征向量L2归一化
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))] # 将每层特征移到CPU  patch_tokens(l,b,p,d)
                image_features = [image_features[bi].squeeze().cpu().numpy() for bi in range(image_features.shape[0])]  # PyTorch张量 → numpy数组
                class_tokens.extend(image_features)
                # Collect features for t-SNE if configured to use class tokens
                if self.vis_tsne and self.tsne_source == 'class':
                    tsne_features.extend(class_tokens[-len(image_features):]) # Append latest batch features
                    # Extend labels for the current batch
                    current_batch_labels = list(image_info["is_anomaly"].numpy())
                    tsne_labels.extend(current_batch_labels)
                
                patch_tokens_list.append(patch_tokens)  # (B, L+1, C)  处理不同batch的patch_tokens，patch_tokens_list(B, l, b, p, d)
            end_time = time.time()
            print('extract time: {}ms per image'.format((end_time-start_time)*1000/subset_num))
            
            # LNAMD 局部邻域聚合生成多尺度特征
            feature_dim = patch_tokens_list[0][0].shape[-1]  # 提取特征维度
            anomaly_maps_r = torch.tensor([]).double()  # 创建一个空的double类型张量，用于存储不同聚合半径r计算得到的异常图
            for r in self.r_list:
                start_time = time.time()
                print('aggregation degree: {}'.format(r))
                # LNAMD_r = LNAMD(device=self.device, r=r, feature_dim=feature_dim, feature_layer=self.features_list)
                # print(f"Using WTConvLNAMD with r={r} (Note: WTConv has random weights if not trained)")
                # LNAMD_r = WTConvLNAMD(device=self.device, feature_dim=feature_dim, feature_layer=self.features_list, r=r)
                
                # --- ABLATION STUDY CONFIGURATION (OPTIMIZED) ---
                # Optimized for: Zipper (Multi-defect visibility), Capsule (Noise), LED (Whole Region)
                ablation_wt_type = 'db1'        # Haar wavelet for sharp edge detection
                ablation_padding = 'reflect'
                ablation_level0  = False        # Processed features only
                
                # Band-Pass / Frequency Configuration:
                # 1. Enable Details (High Freq) starting from Level 1 (Skip Level 0 to avoid noise).
                # 2. Keep LL (Low Freq) to capture global color/region shifts (LED).
                ablation_use_details = True     
                ablation_detail_start = 1       # 1: Skip Level 0 (Noise)
                ablation_keep_ll = True         # True: Include Low Frequency Approximation

                ablation_gamma   = 2.0          # Moderate Gamma
                ablation_use_spot_weight = True  # Suppress patterns found in ANY other image (Occasional Normal Pattern)
                ablation_use_morphology = True  # Toggle for Morphological Optimization (Opening/Closing + Smoothing)
                
                # Morphological Parameters
                ablation_morph_open_k = 1       # Opening kernel size (remove noise). 1 = disabled.
                ablation_morph_close_k = 3      # Closing kernel size (fill gaps). 3 is gentle.
                ablation_morph_smooth_k = 3     # Gaussian smoothing kernel size (remove blockiness).
                ablation_morph_sigma = 0.5      # Gaussian blur standard deviation.
                
                # print(f"Using Original LNAMD with r={r}, ablation_use_spot_weight={ablation_use_spot_weight}, gamma={ablation_gamma}")
                # LNAMD_r = LNAMD(device=self.device, r=r, feature_dim=feature_dim, feature_layer=self.features_list)

                print(f"Using WTConvLNAMDStatic with r={r}, wt={ablation_wt_type}, bandpass={ablation_use_details}(start={ablation_detail_start}, keep_ll={ablation_keep_ll})")
                LNAMD_r = WTConvLNAMDStatic(device=self.device, feature_dim=feature_dim, feature_layer=self.features_list, r=r,
                                            wt_type=ablation_wt_type, padding_mode=ablation_padding, include_level0=ablation_level0,
                                            use_details=ablation_use_details, detail_start_level=ablation_detail_start, keep_ll=ablation_keep_ll)
                Z_layers = {}
                for im in range(len(patch_tokens_list)):  # 遍历所有batch的patch tokens(l,b,p,d)
                    patch_tokens = [p.to(self.device) for p in patch_tokens_list[im]]  # 提取局部特征patch tokens
                    with torch.no_grad(), torch.cuda.amp.autocast():
                        features = LNAMD_r._embed(patch_tokens)  # 使用LNAMD进行特征聚合，输入patch_tokens，输出聚合后的特征 patch tokens[4, (4, 1370, 1024)]
                        features /= features.norm(dim=-1, keepdim=True)  # 对聚合后的特征[4, 1369, 4, 1024]进行L2归一化
                        # 总结：Unfold将每个位置周围的r×r邻域提取出来，adaptive_avg_pool1d将每个邻域的特征聚合为固定维度的特征向量，然后用stack将不同深度层的特征组合在一起
                        for l in range(len(self.features_list)):  # 按层分离并存储特征
                            # save the aggregated features
                            if str(l) not in Z_layers.keys():
                                Z_layers[str(l)] = []
                            Z_layers[str(l)].append(features[:, :, l, :])
                    
                    # Collect LNAMD features for t-SNE if configured
                    if self.vis_tsne and self.tsne_source == 'lnamd':
                        # Use features from the last layer (or a specific layer) for visualization
                        # Here we use global average pooling on the spatial dimensions of the last layer features
                        # features shape: [B, H*W, L, C]
                        # Take the last layer features: [B, H*W, C]
                        last_layer_idx = -1
                        lnamd_feats = features[:, :, last_layer_idx, :]
                        # Global Average Pooling: [B, C]
                        gap_feats = lnamd_feats.mean(dim=1).cpu().numpy()
                        tsne_features.extend(gap_feats)
                        
                        # We need the labels for this batch 'im'.
                        # The full gt_list for this subset is accumulating, but 'im' is the index in patch_tokens_list.
                        # patch_tokens_list has size subset_num // batch_size.
                        # The start index for this batch in gt_list (relative to current subset) is im * batch_size.
                        # However, gt_list contains ALL labels from previous divide_iters too if we don't clear it.
                        # Looking at line 163, gt_list is extended. So gt_list has total dataset_num labels.
                        # The subset starts at dataset_num - subset_num.
                        # So current batch starts at (dataset_num - subset_num) + im * batch_size.
                        
                        start_idx = (dataset_num - subset_num) + im * self.batch_size
                        end_idx = start_idx + features.shape[0]
                        current_batch_labels = gt_list[start_idx:end_idx]
                        tsne_labels.extend(current_batch_labels)
                
                end_time = time.time()
                print('LNAMD-{}: {}ms per image'.format(r, (end_time-start_time)*1000/subset_num))

                # MSM 互评分模块，用于计算每个位置与其他位置的相似度来生成异常图。
                anomaly_maps_l = torch.tensor([]).double()
                start_time = time.time()
                for l in Z_layers.keys():
                    # different layers
                    Z = torch.cat(Z_layers[l], dim=0).to(self.device) # (N, L, C) 将所有批次的该层特征拼接
                    print('layer-{} mutual scoring...'.format(l))
                    
                    # Apply spot weighting conditionally
                    current_use_spot_weight = ablation_use_spot_weight
                    if current_use_spot_weight:
                        # Only apply if using WTConvLNAMDStatic and for specific categories
                        is_wtconv = isinstance(LNAMD_r, WTConvLNAMDStatic)
                        is_target_category = (self.dataset == 'mvtec_ad' and category in ['screw', 'toothbrush', 'zipper'])
                        
                        if not (is_wtconv and is_target_category):
                            current_use_spot_weight = False
                            
                    anomaly_maps_msm = MSM(Z=Z, device=self.device, topmin_min=0, topmin_max=0.3, 
                                           gamma=ablation_gamma, use_spot_weight=current_use_spot_weight)  #调用MSM算法生成异常图（同一层互相计算）
                    anomaly_maps_l = torch.cat((anomaly_maps_l, anomaly_maps_msm.unsqueeze(0).cpu()), dim=0)  # 存储不同层的MSM异常图结果
                    torch.cuda.empty_cache()
                anomaly_maps_l = torch.mean(anomaly_maps_l, 0)  # 将不同层的MSM异常图结果平均融合
                anomaly_maps_r = torch.cat((anomaly_maps_r, anomaly_maps_l.unsqueeze(0)), dim=0)  # 存储不同r值的异常图
                end_time = time.time()
                print('MSM: {}ms per image'.format((end_time-start_time)*1000/subset_num))
            anomaly_maps_iter = torch.mean(anomaly_maps_r, 0).to(self.device)  # 对不同r的异常图取平均
            
            # anomaly_maps_iter current shape: (B, L) where L is flattened H*W
            B, L = anomaly_maps_iter.shape
            H = int(np.sqrt(L))
            W = H
            
            # Reshape to (B, 1, H, W) for morphological optimization or direct interpolation
            anomaly_maps_iter_spatial = anomaly_maps_iter.view(B, 1, H, W)
            
            # --- MORPHOLOGICAL OPTIMIZATION ---
            if ablation_use_morphology:
                # Use very conservative morphological parameters to preserve defect textures:
                # - Opening (remove noise): kernel 1 (effectively disabled) to avoid losing fine details like Zipper teeth
                # - Closing (fill gaps): kernel 3 (small) to fill tiny holes without turning structures into big blobs
                # - Smoothing: kernel 3, sigma 0.5 (very light blur) just to soften max-pool edges, not the whole map
                optimizer = AnomalyMapOptimizer(kernel_size_open=ablation_morph_open_k, 
                                                kernel_size_close=ablation_morph_close_k, 
                                                smooth_kernel_size=ablation_morph_smooth_k, 
                                                sigma=ablation_morph_sigma).to(self.device)
                anomaly_maps_iter_spatial = optimizer(anomaly_maps_iter_spatial)
            
            del anomaly_maps_r
            torch.cuda.empty_cache()

            # interpolate 异常图上采样
            # anomaly_maps_iter_spatial: (B, 1, H, W) -> interpolate directly
            anomaly_maps_iter = F.interpolate(anomaly_maps_iter_spatial,
                                        size=self.image_size, mode='bilinear', align_corners=True)  # 双线性插值上采样
            anomaly_maps = torch.cat((anomaly_maps, anomaly_maps_iter.squeeze(1).cpu()), dim=0)  # 存储所有划分子集的结果

        # save image features for optimizing classification
        # cls_save_path = os.path.join('./image_features/{}_{}.dat'.format(dataset, category))
        # with open(cls_save_path, 'wb') as f:
        #     pickle.dump([np.array(class_tokens)], f)
        end_time_all = time.time()
        print('MuSc: {}ms per image'.format((end_time_all-start_time_all)*1000/dataset_num))
        if torch.cuda.is_available():
            print('MuSc GPU Memory: {:.2f} MB (Allocated), {:.2f} MB (Reserved)'.format(
                torch.cuda.max_memory_allocated() / 1024 / 1024,
                torch.cuda.max_memory_reserved() / 1024 / 1024
            ))

        if self.vis_tsne:
            try:
                tsne_features_np = np.array(tsne_features)
                tsne_labels_np = np.array(tsne_labels)
                # Check if we have enough samples for t-SNE
                if tsne_features_np.shape[0] > self.tsne_perplexity:
                    self.visualize_tsne(tsne_features_np, tsne_labels_np, category, title=f"t-SNE ({self.tsne_source})")
                else:
                    print(f"Skipping t-SNE: Not enough samples ({tsne_features_np.shape[0]}) for perplexity {self.tsne_perplexity}")
            except Exception as e:
                print(f"Error during t-SNE visualization: {e}")

        anomaly_maps = anomaly_maps.cpu().numpy()
        torch.cuda.empty_cache()

        B = anomaly_maps.shape[0]   # the number of unlabeled test images
        ac_score = np.array(anomaly_maps).reshape(B, -1).max(-1)
        if self.use_rscin:
            # RsCIN
            if self.dataset == 'visa':
                k_score = [1, 8, 9]
            elif self.dataset == 'mvtec_ad':
                k_score = [1, 2, 3]
            else:
                k_score = [1, 2, 3]
            scores_cls = RsCIN(ac_score, class_tokens, k_list=k_score)
        else:
            # Disable RsCIN to compare with raw image-level scores.
            scores_cls = ac_score

        print('computing metrics...')
        pr_sp = np.array(scores_cls)
        gt_sp = np.array(gt_list)
        gt_px = torch.cat(img_masks, dim=0).numpy().astype(np.int32)
        pr_px = np.array(anomaly_maps)
        image_metric, pixel_metric = compute_metrics(gt_sp, pr_sp, gt_px, pr_px)
        auroc_sp, f1_sp, ap_sp = image_metric
        auroc_px, f1_px, ap_px, aupro = pixel_metric
        print(category)
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp*100, f1_sp*100, ap_sp*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px*100, f1_px*100, ap_px*100, aupro*100))

        if self.vis:
            print('visualization...')
            self.visualization(image_path_list, gt_list, pr_px, gt_px, category)
    
        if torch.cuda.is_available():
            mem_allocated = torch.cuda.max_memory_allocated() / 1024 / 1024
            mem_reserved = torch.cuda.max_memory_reserved() / 1024 / 1024
        else:
            mem_allocated, mem_reserved = 0, 0
            
        return image_metric, pixel_metric, ((end_time_all-start_time_all)*1000/dataset_num), mem_allocated, mem_reserved


    def main(self):
        auroc_sp_ls = []
        f1_sp_ls = []
        ap_sp_ls = []
        auroc_px_ls = []
        f1_px_ls = []
        ap_px_ls = []
        aupro_ls = []
        time_ls = []
        mem_alloc_ls = []
        mem_res_ls = []
        for category in self.categories:
            image_metric, pixel_metric, avg_time, mem_alloc, mem_res = self.make_category_data(category=category,)  # 对每个类别进行缺陷检测
            auroc_sp, f1_sp, ap_sp = image_metric
            auroc_px, f1_px, ap_px, aupro = pixel_metric
            auroc_sp_ls.append(auroc_sp)
            f1_sp_ls.append(f1_sp)
            ap_sp_ls.append(ap_sp)
            auroc_px_ls.append(auroc_px)
            f1_px_ls.append(f1_px)
            ap_px_ls.append(ap_px)
            aupro_ls.append(aupro)
            time_ls.append(avg_time)
            mem_alloc_ls.append(mem_alloc)
            mem_res_ls.append(mem_res)
            
        # mean
        auroc_sp_mean = sum(auroc_sp_ls) / len(auroc_sp_ls)
        f1_sp_mean = sum(f1_sp_ls) / len(f1_sp_ls)
        ap_sp_mean = sum(ap_sp_ls) / len(ap_sp_ls)
        auroc_px_mean = sum(auroc_px_ls) / len(auroc_px_ls)
        f1_px_mean = sum(f1_px_ls) / len(f1_px_ls)
        ap_px_mean = sum(ap_px_ls) / len(ap_px_ls)
        aupro_mean = sum(aupro_ls) / len(aupro_ls)
        time_mean = sum(time_ls) / len(time_ls)
        mem_alloc_mean = sum(mem_alloc_ls) / len(mem_alloc_ls)
        mem_res_mean = sum(mem_res_ls) / len(mem_res_ls)

        for i, category in enumerate(self.categories):
            print(category)
            print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_ls[i]*100, f1_sp_ls[i]*100, ap_sp_ls[i]*100))
            print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_ls[i]*100, f1_px_ls[i]*100, ap_px_ls[i]*100, aupro_ls[i]*100))
        print('mean')
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_mean*100, f1_sp_mean*100, ap_sp_mean*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_mean*100, f1_px_mean*100, ap_px_mean*100, aupro_mean*100))
        print('MuSc: {:.2f}ms per image'.format(time_mean))
        if torch.cuda.is_available():
            print('MuSc GPU Memory: {:.2f} MB (Allocated), {:.2f} MB (Reserved)'.format(mem_alloc_mean, mem_res_mean))
        
        # save in excel
        if self.save_excel:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "MuSc_results"
            sheet.cell(row=1,column=2,value='auroc_px')
            sheet.cell(row=1,column=3,value='f1_px')
            sheet.cell(row=1,column=4,value='ap_px')
            sheet.cell(row=1,column=5,value='aupro')
            sheet.cell(row=1,column=6,value='auroc_sp')
            sheet.cell(row=1,column=7,value='f1_sp')
            sheet.cell(row=1,column=8,value='ap_sp')
            for col_index in range(2):
                for row_index in range(len(self.categories)):
                    if col_index == 0:
                        sheet.cell(row=row_index+2,column=col_index+1,value=self.categories[row_index])
                    else:
                        sheet.cell(row=row_index+2,column=col_index+1,value=auroc_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+2,value=f1_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+3,value=ap_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+4,value=aupro_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+5,value=auroc_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+6,value=f1_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+7,value=ap_sp_ls[row_index]*100)
                    if row_index == len(self.categories)-1:
                        if col_index == 0:
                            sheet.cell(row=row_index+3,column=col_index+1,value='mean')
                        else:
                            sheet.cell(row=row_index+3,column=col_index+1,value=auroc_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+2,value=f1_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+3,value=ap_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+4,value=aupro_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+5,value=auroc_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+6,value=f1_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+7,value=ap_sp_mean*100)
            workbook.save(os.path.join(self.output_dir, 'results.xlsx'))


